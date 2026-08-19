#!/usr/bin/env python3
"""MLBMA daily board build (in-session per SPEC v1.1 / v1.0 / PARAMS / TT park addendum).
Inputs assumed in cwd: wb.xlsx, games.json, lineups.json, splits.json, sp_hands.json,
odds_out/*, model_a.json, model_b.json, props_page.json, weather.json.
Outputs: projections/DATE.csv, candidates/DATE.csv, board.json (for emission).
"""
import json, csv, math, statistics, re, unicodedata, sys, os
from collections import defaultdict
import openpyxl
sys.path.insert(0, '.')
from pricing import am_to_prob, prob_to_am, novig_two_way, is_corrupt_am

import datetime as _dt
DATE = os.environ.get('MLBMA_DATE') or _dt.date.today().isoformat()
TS = os.environ.get('MLBMA_TS') or _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:00Z')
ACTIONABLE = {'draftkings', 'fanduel', 'williamhill_us', 'betmgm', 'betrivers', 'pinnacle', 'lowvig', 'hardrockbet', 'bet365'}
ALIAS = {'ARI': 'AZ', 'CHW': 'CWS', 'OAK': 'ATH', 'WAS': 'WSH', 'KCR': 'KC', 'SDP': 'SD', 'SFG': 'SF', 'TBR': 'TB', 'WSN': 'WSH'}
def A(t): return ALIAS.get(t, t)
def base(t): return t.split('#')[0]   # DH-safe team key: 'STL#1' -> 'STL' for stat lookups; suffix only on DH game 1
def norm(s):
    s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode().lower()
    s = re.sub(r'\b(jr|sr|ii|iii|iv)\.?$', '', s.strip()).replace('.', '').replace("'", '').strip()
    return re.sub(r'\s+', ' ', s)

# ---------------- constants (PARAMS SNAPSHOT) ----------------
SLOT_PA = {1: 4.507, 2: 4.406, 3: 4.309, 4: 4.199, 5: 4.053, 6: 3.901, 7: 3.778, 8: 3.599, 9: 3.457}
PA_MULT = {1: 1.12031, 2: 1.09520, 3: 1.07109, 4: 1.04375, 5: 1.00746, 6: 0.96967, 7: 0.93910, 8: 0.89461, 9: 0.85931}
SLOT_PP = {1: 4.76, 2: 3.77, 3: 2.81, 4: 1.73, 5: 0.30, 6: -1.20, 7: -2.41, 8: -4.17, 9: -5.57}
LEAGUE_GAP = {'L': .1321, 'R': .0415, 'S': .0036}
TB_PARK_PP = {'Coors Field': 8.92, 'Chase Field': 2.08, 'Target Field': 1.73, 'Nationals Park': 1.72, 'Oriole Park at Camden Yards': 1.49,
              'Kauffman Stadium': 1.31, 'PNC Park': 1.04, 'loanDepot park': .84, 'Rogers Centre': .78, 'Great American Ball Park': .64,
              'Angel Stadium': .25, 'Fenway Park': .19, 'Citizens Bank Park': .04, 'Progressive Field': -.07, 'Truist Park': -.47,
              'Busch Stadium': -.53, 'Oracle Park': -.76, 'Comerica Park': -1.03, 'Citi Field': -1.48, 'Yankee Stadium': -1.49,
              'Wrigley Field': -2.28, 'American Family Field': -2.47, 'Petco Park': -2.73, 'T-Mobile Park': -3.45, 'Globe Life Field': -3.51}
PARK_TEMP = {'George M. Steinbrenner Field': 83.7, 'Sutter Health Park': 79.6, 'Truist Park': 79.1, 'Busch Stadium': 78.0,
             'Oriole Park at Camden Yards': 78.0, 'Kauffman Stadium': 77.6, 'Great American Ball Park': 76.9, 'Citizens Bank Park': 75.9,
             'Coors Field': 75.1, 'Angel Stadium': 74.7, 'Nationals Park': 74.7, 'PNC Park': 73.4, 'Target Field': 72.8, 'Citi Field': 72.3,
             'Comerica Park': 71.7, 'Dodger Stadium': 71.6, 'UNIQLO Field at Dodger Stadium': 71.6, 'Yankee Stadium': 71.4, 'Rate Field': 71.3,
             'Progressive Field': 71.2, 'Fenway Park': 70.3, 'T-Mobile Park': 69.8, 'Petco Park': 68.7, 'Wrigley Field': 66.8, 'Oracle Park': 63.0}
WIND_MULT = {'Wrigley Field': 2.44, 'Oriole Park at Camden Yards': 1.75, 'Rogers Centre': 1.59, 'Rate Field': 1.58, 'PNC Park': 1.27, 'Truist Park': 1.22,
             'Comerica Park': 1.12, 'Great American Ball Park': 1.06, 'Fenway Park': .90, 'Target Field': .85, 'Citi Field': .79, 'Busch Stadium': .75,
             'Citizens Bank Park': .45, 'Sutter Health Park': 1.0, 'Oracle Park': 0.0}
TB_WIND_PP = {('OUT', 1): 1.43, ('OUT', 2): 1.26, ('IN', 1): -.72, ('IN', 2): -1.28, ('X', 1): 1.23, ('X', 2): 1.73}
TT_WIND_RUNS = {('OUT', 1): .42, ('OUT', 2): .56, ('IN', 1): -.09, ('IN', 2): -.89, ('X', 1): .38, ('X', 2): .655}
TT_PARK = {'COL': 1.28, 'CIN': 1.12, 'BOS': 1.08, 'PHI': 1.06, 'AZ': 1.05, 'ATH': 1.05, 'CWS': 1.04, 'NYY': 1.04, 'CHC': 1.03, 'BAL': 1.02, 'TEX': 1.02,
           'ATL': 1.02, 'KC': 1.02, 'TOR': 1.01, 'MIN': 1.00, 'WSH': 1.00, 'LAA': .99, 'HOU': .99, 'MIL': .98, 'STL': .98, 'LAD': .98, 'NYM': .97,
           'DET': .97, 'CLE': .97, 'MIA': .96, 'TB': .95, 'SD': .95, 'PIT': .94, 'SF': .94, 'SEA': .92}
DOMES = {'Tropicana Field', 'Rogers Centre', 'Daikin Park', 'Minute Maid Park', 'Globe Life Field', 'loanDepot park', 'Chase Field', 'T-Mobile Park', 'American Family Field'}

def z(vals):
    xs = [v for v in vals if v is not None]
    if len(xs) < 2: return [0.0 for _ in vals]
    m = statistics.mean(xs); s = statistics.pstdev(xs) or 1e-9
    return [((v - m) / s) if v is not None else 0.0 for v in vals]
def zmap(d):
    ks = list(d); zs = z([d[k] for k in ks]); return dict(zip(ks, zs))
def clip(x, lo, hi): return max(lo, min(hi, x))
def shrink(x, n, mean): return (n * x + 6 * mean) / (n + 6)
def fnum(v):
    try: return float(v)
    except: return None

# ---------------- load ----------------
games = json.load(open('games.json')); lineups = json.load(open('lineups.json')); splits = json.load(open('splits.json'))
for _g in games:
    _suf = '#1' if (_g.get('dh') in ('S','Y') and _g.get('gn') == 1) else ''
    _g['away_k'] = _g['away'] + _suf; _g['home_k'] = _g['home'] + _suf; _g['mkey'] = f"{_g['away']}@{_g['home']}" + ('#'+str(_g['gn']) if _g.get('dh') in ('S','Y') else '')
sp_hands = {int(k): v for k, v in json.load(open('sp_hands.json')).items()}
model_a = json.load(open('model_a.json')); model_b = json.load(open('model_b.json')); props_page = json.load(open('props_page.json'))
weather = json.load(open('weather.json'))
wb = openpyxl.load_workbook('wb.xlsx', data_only=True, read_only=True)

# hitters tab by mlbamid
H = {}
for r in wb['Hitters, 26'].iter_rows(min_row=3, values_only=True):
    if not r[0] or not isinstance(r[3], (int, float)) or not r[48]: continue
    H[int(r[48])] = {'name': r[0], 'team': A(r[1]), 'PA': r[3], 'R': r[6], 'RBI': r[7], 'xSLG': r[14], 'OPS': r[15], 'ISO': r[16], 'xwOBA': r[18],
                     'wRC': r[19], 'K': r[20], 'BB': r[21], 'Chase': r[24], 'Pull': r[32], 'HH95': r[34], 'BRL': r[35], 'H': r[40]}
# SP tab (PROBABLES has all 30, full 93 cols)
SP = {}
for r in wb[[t for t in wb.sheetnames if t.startswith('PROBABLES')][0]].iter_rows(min_row=3, values_only=True):
    if not r[3]: continue
    g = lambda i: fnum(r[i - 1])
    _rec = {'team': A(r[2]), 'name': r[3], 'IP': g(5), 'ERA': g(6), 'WHIP': g(7), 'xERA': g(8), 'SIERA': g(9), 'Stuff': g(14), 'K%': g(15), 'BB%': g(16),
        'KBB': g(17), 'SwStr': g(18), 'CSW': g(19), 'Ball': g(20), 'FStrk': g(21), 'Chase': g(22), 'BAA': g(25), 'BABIP': g(26), 'GB': g(27), 'FB': g(29),
        'HRFB': g(32), 'HH': g(33), 'BRL': g(35), 'GS': g(38), 'K': g(39), 'BB': g(40),
        'L': {'TBF': g(45), 'xFIP': g(46), 'BAA': g(47), 'OPS': g(48), 'BABIP': g(49), 'WHIP': g(50), 'GB': g(53), 'FB': g(54), 'HRFB': g(56), 'HH': g(57)},
        'R': {'TBF': g(60), 'xFIP': g(61), 'BAA': g(62), 'OPS': g(63), 'BABIP': g(64), 'WHIP': g(65), 'GB': g(68), 'FB': g(69), 'HRFB': g(71), 'HH': g(72)},
        'K/GS': g(75), 'BB/GS': g(76), 'OUT/GS': g(77), 'L30_xERA': g(91), 'L30_SIERA': g(92), 'L30_KBB': g(93)}
    SP[norm(r[3])] = _rec
    _bare = norm(re.sub(r'\([^)]*\)', '', str(r[3])))   # 'Ethan Pecko (AAA)' -> 'ethan pecko'; AAA/AA call-ups carry the level in the workbook name
    if _bare and _bare not in SP: SP[_bare] = _rec
# bullpens
PEN = {}
for r in wb['Bullpens'].iter_rows(min_row=3, max_col=34, values_only=True):
    if r[2]: PEN[A(r[2])] = {'ERA2wk': fnum(r[5]), 'BAA2wk': fnum(r[8]), 'ERAytd': fnum(r[20]), 'BAAytd': fnum(r[23])}
# team offense YTD (fallback + league avgs)
TMO = {}
for r in wb['Tm O YTD'].iter_rows(min_row=3, max_col=34, values_only=True):
    if r[2]: TMO[A(r[2])] = {'OPS': fnum(r[12]), 'K': fnum(r[14]), 'BB': fnum(r[15]), 'Chase': fnum(r[17]), 'xwOBA': fnum(r[31])}
LG = {k: statistics.mean(v[k] for v in TMO.values() if v[k] is not None) for k in ('OPS', 'K', 'BB', 'Chase', 'xwOBA')}
# ITT tab
ITT = {}
for r in wb['ITT!'].iter_rows(min_row=3, max_col=6, values_only=True):
    if r[1] and isinstance(r[2], (int, float)):
        _t = str(r[1]); _k = A(_t[1:]) + '#1' if _t[0] in 'yz' and _t[1:].isupper() else A(_t)
        ITT[_k] = float(r[2])
SPRANK = {}
for r in wb['SP Ranks+Ks'].iter_rows(min_row=3, max_col=6, values_only=True):
    if r[2] and r[2] != 'BULLPEN': SPRANK[norm(r[2])] = r[5]

# ---------------- odds ----------------
raw_props = json.load(open(f'odds_out/raw_props_{DATE}.json'))
events = json.load(open(f'odds_out/events_{DATE}.json'))
gl = {r['event_id']: r for r in csv.DictReader(open(f'odds_out/game_lines_{DATE}.csv'))}
alt_tot = json.load(open('odds_out/alt_totals.json'))
def gk(away, home, gn=None): return f"{'CHW' if away=='CWS' else away}@{'CHW' if home=='CWS' else home}" + (f'#{gn}' if gn else '')
# per-event per-book quotes: quotes[eid][market][(player,point)] = list of (book, over, under)
quotes = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
for eid, data in raw_props.items():
    for bk in data.get('bookmakers', []):
        for mk in bk.get('markets', []):
            pairs = defaultdict(dict)
            for o in mk.get('outcomes', []):
                pairs[(o.get('description'), o.get('point'))][o.get('name')] = o.get('price')
            for k, v in pairs.items():
                quotes[eid][mk['key']][k].append((bk['key'], v.get('Over'), v.get('Under')))
def clean(xs): return [x for x in xs if x is not None and not is_corrupt_am(x)]
def med_prob(prices):
    p = clean(prices); return statistics.median(am_to_prob(x) for x in p) if p else None
def price_summary(rows):
    """rows: list of (book, over, under) -> med over/under am, novig over prob, best actionable over (price, book)."""
    ov = [r[1] for r in rows]; un = [r[2] for r in rows]
    po, pu = med_prob(ov), med_prob(un)
    nv = novig_two_way(po, pu) if (po is not None and pu is not None) else po
    act = [(r[1], r[0]) for r in rows if r[0] in ACTIONABLE and r[1] is not None and not is_corrupt_am(r[1])]
    best = max(act, key=lambda t: am_to_prob(t[0]) * -1) if act else (None, None)  # highest payout = lowest implied prob
    med_o = prob_to_am(po) if po is not None else None
    med_u = prob_to_am(pu) if pu is not None else None
    return {'med_over': med_o, 'med_under': med_u, 'novig_over': nv, 'best_over': best[0], 'best_book': best[1], 'n': len(rows)}
def modal_point(mkt_q):
    pts = defaultdict(int)
    for (pl, pt), rows in mkt_q.items(): pts[pt] += len(rows)
    return max(pts, key=pts.get) if pts else None
def event_for(away, home, gn=None):
    from pull_odds import team_abbr
    c = [ev for ev in events if team_abbr(ev['away_team']) == ('CHW' if away == 'CWS' else away) and team_abbr(ev['home_team']) == ('CHW' if home == 'CWS' else home)]
    c.sort(key=lambda e: e['commence_time'])
    if not c: return None
    if gn and len(c) >= gn: return c[gn - 1]['id']
    return c[0]['id']
def player_quotes(eid, market, pname, point=None):
    q = quotes[eid][market]; out = {}
    for (pl, pt), rows in q.items():
        if pl and norm(pl) == norm(pname) and (point is None or pt == point): out[pt] = rows
    return out
def payout(am): return am if am > 0 else 100 * 100 / (-am)

# ---------------- weather / park per game ----------------
gm = {}
for g in games:
    gp = str(g['game_pk']); w = weather[gp]
    venue = g['venue']; dome = venue in DOMES or w.get('dome')
    tb_wx = 0.0; tt_wx = 0.0; wbin = 0; wdir = w.get('dir')
    if not dome:
        mph = w.get('mph', 0) or 0
        wbin = 1 if 10 <= mph < 15 else (2 if mph >= 15 else 0)
        if wbin and wdir in ('OUT', 'IN', 'X'):
            mult = WIND_MULT.get(venue, 1.0) if wdir == 'OUT' else 1.0
            tb_wx += TB_WIND_PP[(wdir, wbin)] * mult
            tt_wx += 0.714 * TT_WIND_RUNS[(wdir, wbin)] * mult
        pm = PARK_TEMP.get(venue, 72.0)
        tb_wx += 0.1215 * (w['temp'] - pm)
        tt_wx += 0.714 * 0.0355 * (w['temp'] - pm)
    gm[gp] = {'g': g, 'dome': dome, 'tb_wx': tb_wx, 'tt_wx': tt_wx, 'wbin': wbin, 'wdir': wdir, 'w': w, 'venue': venue,
              'park_pp': TB_PARK_PP.get(venue, 0.0), 'park_tt': TT_PARK[g['home']], 'eid': event_for(g['away'], g['home'], g.get('gn') if g.get('dh') in ('S','Y') else None)}

# ---------------- pitcher table (per starter, incl. opener protocol) ----------------
OPENERS = {'Lake Bachar'}  # PIT listed BULLPEN in workbook; Bachar 8.5-out profile
P = {}
for g in games:
    gp = str(g['game_pk'])
    for side in ('away', 'home'):
        nm = g[side + '_sp']; pid = g[side + '_sp_id']; team = g[side + '_k']; opp = g[('home' if side == 'away' else 'away') + '_k']
        hand = sp_hands[pid][1]
        sp = SP.get(norm(nm))
        pp = props_page.get(norm(nm), {})
        # outgs_adj = props-page OUTS proj (interp 1); fallback shrunk OUT/GS later
        P[pid] = {'id': pid, 'name': nm, 'team': team, 'opp': opp, 'hand': hand, 'gp': gp, 'side': side, 'sp': sp, 'pp': pp,
                  'opener': nm in OPENERS or sp is None, 'outs_pp': pp.get('OUTS', {}).get('proj')}
# slate means for shrink
def spv(k): return [p['sp'][k] for p in P.values() if p['sp'] and p['sp'].get(k) is not None]
mean_outgs = statistics.mean(spv('OUT/GS')); mean_kgs = statistics.mean(spv('K/GS')); mean_bbgs = statistics.mean(spv('BB/GS'))
for p in P.values():
    sp = p['sp']
    if sp and sp['GS']:
        p['out_shr'] = shrink(sp['OUT/GS'], sp['GS'], mean_outgs)
        p['outgs_adj'] = p['outs_pp'] if p['outs_pp'] else p['out_shr']
    else:
        p['out_shr'] = mean_outgs; p['outgs_adj'] = p['outs_pp'] or 9.0
    p['w_sp'] = clip(0.081 + 0.03272 * p['outgs_adj'], 0.30, 0.80)
    p['smGS'] = bool(sp and sp['GS'] is not None and sp['GS'] < 8)

# ---------------- lineups: hitter rows ----------------
def hand_ops(pid, vs_hand):
    """EB-shrunk split OPS vs pitcher hand (k=450 toward overall + league gap/2)."""
    s = splits.get(str(pid)) or {}
    hh = H.get(pid)
    overall = (hh['OPS'] if hh and hh['OPS'] else (s.get('season') or {}).get('OPS'))
    if overall is None: return None, None
    if hh is None:  # not on the workbook (tiny sample): shrink the raw season OPS toward league k=450 first
        pa0 = (s.get('season') or {}).get('PA', 0) or 0
        overall = (pa0 * overall + 450 * LG['OPS']) / (pa0 + 450)
    bats = s.get('bats') or 'R'
    key = 'vl' if vs_hand == 'L' else 'vr'
    opp_hand = (bats == 'R' and vs_hand == 'L') or (bats == 'L' and vs_hand == 'R') or bats == 'S'
    gap = LEAGUE_GAP.get(bats, .04)
    expected = overall + (gap / 2 if opp_hand else -gap / 2)
    if bats == 'S': expected = overall
    sp = s.get(key)
    if sp and sp['PA']:
        return (sp['PA'] * sp['OPS'] + 450 * expected) / (sp['PA'] + 450), overall
    return expected, overall

L = []  # hitter rows
for g in games:
    gp = str(g['game_pk'])
    for side in ('away', 'home'):
        team = g[side + '_k']; opp = g[('home' if side == 'away' else 'away') + '_k']
        opp_pid = g[('home' if side == 'away' else 'away') + '_sp_id']
        lu = lineups[gp][side]
        for pl in lu['players']:
            pid = pl['id']; hh = H.get(pid); s = splits.get(str(pid), {})
            so, ov = hand_ops(pid, P[opp_pid]['hand'])
            L.append({'id': pid, 'name': pl['name'], 'team': team, 'opp': opp, 'gp': gp, 'slot': pl['slot'], 'src': lu['src'], 'hh': hh,
                      'bats': s.get('bats'), 'PA': (hh['PA'] if hh else (s.get('season') or {}).get('PA', 0)) or 0,
                      'split_ops': so, 'ovr_ops': ov, 'opp_pid': opp_pid, 'no_wb': hh is None})
byteam = defaultdict(list)
for h in L: byteam[h['team']].append(h)

def lw(team, key, hand_adj=False):
    """PA-weighted (SLOT_PA) lineup aggregate of hitters-tab stat `key`; renormalize over hitters that have it."""
    num = den = 0.0
    for h in byteam[team]:
        if h['hh'] and h['hh'].get(key) is not None:
            w = SLOT_PA[h['slot']]; num += w * h['hh'][key]; den += w
    return num / den if den else None
def lw_ops_vs_hand(team):
    num = den = 0.0
    for h in byteam[team]:
        if h['split_ops'] is not None:
            w = SLOT_PA[h['slot']]; num += w * h['split_ops']; den += w
    return num / den if den else TMO[base(team)]['OPS']
def lr_share(team):
    """share of lineup PA from L-side bats vs the opposing SP hand (for hand-weighted OPS allowed)."""
    opp_hand = P[byteam[team][0]['opp_pid']]['hand'] if byteam[team] else 'R'
    l = r = 0.0
    for h in byteam[team]:
        w = SLOT_PA[h['slot']]; b = h['bats'] or 'R'
        if b == 'S': b = 'L' if opp_hand == 'R' else 'R'
        if b == 'L': l += w
        else: r += w
    return (l / (l + r), r / (l + r)) if (l + r) else (0.5, 0.5)

# ---------------- opponent-adjusted pitcher terms ----------------
for p in P.values():
    t = p['opp']
    p['opp_K'] = lw(t, 'K') or TMO[base(t)]['K']; p['opp_BB'] = lw(t, 'BB') or TMO[base(t)]['BB']; p['opp_Chase'] = lw(t, 'Chase') or TMO[base(t)]['Chase']
    p['opp_xwOBA'] = lw(t, 'xwOBA') or TMO[base(t)]['xwOBA']
    p['opp_adj_K'] = clip(p['opp_K'] / LG['K'], .8, 1.2); p['opp_adj_BB'] = clip(p['opp_BB'] / LG['BB'], .8, 1.2)
    p['opp_adj_ER'] = clip(p['opp_xwOBA'] / LG['xwOBA'], .8, 1.2)
    p['team_agg'] = not byteam[t]

# ---------------- K / BB / ER models ----------------
KP = [p for p in P.values() if p['sp'] and not p['opener'] and p['sp'].get('K/GS') is not None and p['sp'].get('BB/GS') is not None and p['sp'].get('SIERA') is not None]
def zk(key, src=lambda p: p['sp']): return dict(zip([p['id'] for p in KP], z([src(p).get(key) if src(p) else None for p in KP])))
def market_line(p, mkt, fallback_key):
    q = player_quotes(gm[p['gp']]['eid'], mkt, p['name'])
    if q:
        pt = max(q, key=lambda k: len(q[k]))
        return pt, price_summary(q[pt]), q
    pl = p['pp'].get(fallback_key, {}).get('line')
    return pl, None, {}
for p in KP:
    sp = p['sp']
    p['proj_K'] = shrink(sp['K/GS'], sp['GS'], mean_kgs) * (p['outgs_adj'] / p['out_shr']) * p['opp_adj_K']
    p['proj_BB'] = shrink(sp['BB/GS'], sp['GS'], mean_bbgs) * (p['outgs_adj'] / p['out_shr']) * p['opp_adj_BB']
    skill = shrink(.45 * sp['SIERA'] + .35 * sp['xERA'] + .20 * sp['ERA'], sp['GS'], statistics.mean(.45 * q['SIERA'] + .35 * q['xERA'] + .20 * q['ERA'] for q in (x['sp'] for x in KP)))
    rec = clip(sp['L30_xERA'] / sp['xERA'], .9, 1.1) if (sp['L30_xERA'] and sp['xERA']) else 1.0
    p['recency'] = rec; p['rec_missing'] = not sp['L30_xERA']
    p['proj_ER'] = skill / 27 * p['outgs_adj'] * p['opp_adj_ER'] * rec
    p['K_line'], p['K_px'], _ = market_line(p, 'pitcher_strikeouts', 'K')
    p['BB_line'], p['BB_px'], _ = market_line(p, 'pitcher_walks', 'BB')
    p['ER_line'], p['ER_px'], _ = market_line(p, 'pitcher_earned_runs', 'ER')
# z terms
zS = {k: zk(k) for k in ('SwStr', 'CSW', 'Stuff', 'K%', 'Chase', 'BB%', 'Ball', 'FStrk', 'L30_KBB')}
zgapK = dict(zip([p['id'] for p in KP], z([(p['proj_K'] - p['K_line']) if p['K_line'] is not None else None for p in KP])))
zgapBB = dict(zip([p['id'] for p in KP], z([(p['proj_BB'] - p['BB_line']) if p['BB_line'] is not None else None for p in KP])))
zoppK = dict(zip([p['id'] for p in KP], z([p['opp_K'] for p in KP])))
zoppBB = dict(zip([p['id'] for p in KP], z([p['opp_BB'] for p in KP])))
zoppCh = dict(zip([p['id'] for p in KP], z([p['opp_Chase'] for p in KP])))
zout = dict(zip([p['id'] for p in KP], z([p['outgs_adj'] for p in KP])))
for p in KP:
    i = p['id']
    p['K_score'] = 0.40 * (.34 * zS['SwStr'][i] + .24 * zS['CSW'][i] + .18 * zS['Stuff'][i] + .16 * zS['K%'][i] + .08 * zS['Chase'][i]) \
        + 0.20 * zgapK[i] + 0.15 * zoppK[i] + 0.05 * zS['L30_KBB'][i] + 0.20 * zout[i]
    p['BB_score'] = 0.60 * (.33 * zS['BB%'][i] + .30 * zS['Ball'][i] + .22 * (-zS['FStrk'][i]) + .15 * (-zS['Chase'][i])) \
        + 0.20 * zgapBB[i] + 0.15 * (.67 * zoppBB[i] + .33 * (-zoppCh[i])) + 0.05 * (-zS['L30_KBB'][i])
    p['ER_dist'] = (p['proj_ER'] - p['ER_line']) if p['ER_line'] is not None else None

# ---------------- TT 2.0 ----------------
TT = {}
for g in games:
    gp = str(g['game_pk'])
    for side in ('away', 'home'):
        team = g[side + '_k']; oside = 'home' if side == 'away' else 'away'; osp = P[g[oside + '_sp_id']]
        TT[team] = {'gp': gp, 'opp': g[oside + '_k'], 'osp': osp, 'home': side == 'home'}
# pitcher_vuln components (opposing SP vs THIS lineup)
def sp_hand_wt(sp, key, team):
    if not sp: return None
    ls, rs = lr_share(team)
    lv, rv = sp['L'].get(key), sp['R'].get(key)
    if lv is None or rv is None: return sp.get(key)
    return ls * lv + rs * rv
teams = list(TT)
def tz(vals): return dict(zip(teams, z(vals)))
v_skill = tz([(.5 * TT[t]['osp']['sp']['xERA'] + .5 * TT[t]['osp']['sp']['SIERA']) if TT[t]['osp']['sp'] else None for t in teams])
v_ops = tz([sp_hand_wt(TT[t]['osp']['sp'], 'OPS', t) for t in teams])
v_pow = tz([(.6 * (sp_hand_wt(TT[t]['osp']['sp'], 'HRFB', t) or 0) + .4 * (TT[t]['osp']['sp']['BRL'] or 0)) if TT[t]['osp']['sp'] else None for t in teams])
opp_off = tz([lw_ops_vs_hand(t) for t in teams])
pen_q = tz([.6 * PEN[base(TT[t]['opp'])]['ERA2wk'] + .4 * PEN[base(TT[t]['opp'])]['ERAytd'] for t in teams])
osp_out = tz([TT[t]['osp']['outgs_adj'] for t in teams])
park_z = tz([gm[TT[t]['gp']]['park_tt'] for t in teams])
for t in teams:
    d = TT[t]; bp = d['osp']['sp'] is None  # bullpen game: pitcher_vuln from pen quality only
    pv = (.40 * v_skill[t] + .35 * v_ops[t] + .25 * v_pow[t]) if not bp else pen_q[t]
    d['pitcher_vuln'] = pv; d['opp_offense'] = opp_off[t]
    d['pen_exp'] = .60 * pen_q[t] + .40 * (-osp_out[t]); d['park_z'] = park_z[t]; d['wx'] = gm[d['gp']]['tt_wx']
    d['score'] = .376 * pv + .301 * opp_off[t] + .215 * d['pen_exp'] + .1075 * park_z[t] + d['wx']
    d['bp'] = bp
    # market TT
    eid = gm[d['gp']]['eid']; q = quotes[eid]['team_totals']
    full = [ev for ev in events if ev['id'] == eid][0]
    tname = full['home_team'] if d['home'] else full['away_team']
    tq = {pt: rows for (pl, pt), rows in q.items() if pl == tname}
    d['tq'] = tq
    if tq:
        main = max(tq, key=lambda k: len(tq[k])); d['line_main'] = main
        side = 'Over' if d['score'] > 0 else 'Under'
        bet_line = main
        if main % 1 == 0: bet_line = main - 0.5 if side == 'Over' else main + 0.5
        d['side'] = side; d['line'] = bet_line
        rows = tq.get(bet_line, [])
        ps = price_summary(rows) if rows else None
        d['px'] = ps
        if ps:
            d['price'] = ps['med_over'] if side == 'Over' else ps['med_under']
            d['best'] = ps['best_over'] if side == 'Over' else max([(r[2], r[0]) for r in rows if r[0] in ACTIONABLE and r[2] is not None and not is_corrupt_am(r[2])], key=lambda x: -am_to_prob(x[0]), default=(None, None))[0]
            d['imp'] = ps['novig_over'] if side == 'Over' else (1 - ps['novig_over'] if ps['novig_over'] else None)
        else:
            d['price'] = None; d['best'] = None; d['imp'] = None
    else:
        d['side'] = 'Over' if d['score'] > 0 else 'Under'; d['line'] = None; d['price'] = None; d['best'] = None; d['imp'] = None; d['line_main'] = None
ttz = tz([TT[t]['score'] for t in teams])
itt_z = tz([ITT.get(t) for t in teams])

# ---------------- TB 2.0 + HRR ----------------
hit_pop = [h for h in L]
def hz(key): return z([h['hh'][key] if h['hh'] else None for h in hit_pop])
Q = {k: hz(k) for k in ('wRC', 'xwOBA', 'ISO', 'BRL', 'HH95', 'xSLG', 'Pull', 'K')}
zsplit = z([h['split_ops'] for h in hit_pop]); zovr = z([h['ovr_ops'] for h in hit_pop])
# SP composite vs batter hand (negated z per spec-literal, deferred direction question)
def sp_hand_stat(sp, hand, key):
    if not sp: return None
    v = sp[hand].get(key)
    return v if v is not None else sp.get(key)
def fbgb(sp, hand):
    if not sp: return None
    fb = sp[hand].get('FB') or sp.get('FB'); gb = sp[hand].get('GB') or sp.get('GB')
    return (fb - gb) if (fb is not None and gb is not None) else None
comp_keys = ['xFIP', 'WHIP', 'HH', 'HRFB', 'FBGB', 'BAA', 'BABIP', 'OPS']
comp_vals = {k: [] for k in comp_keys}
for h in hit_pop:
    sp = P[h['opp_pid']]['sp']; hand = 'L' if (h['bats'] == 'L' or (h['bats'] == 'S' and P[h['opp_pid']]['hand'] == 'R')) else 'R'
    for k in comp_keys:
        comp_vals[k].append(fbgb(sp, hand) if k == 'FBGB' else sp_hand_stat(sp, hand, k))
comp_z = {k: z(v) for k, v in comp_vals.items()}
pen_vals = [.6 * PEN[base(h['opp'])]['ERA2wk'] + .4 * PEN[base(h['opp'])]['ERAytd'] for h in hit_pop]
pen_z = z(pen_vals)
pen_baa_z = z([.6 * PEN[base(h['opp'])]['BAA2wk'] + .4 * PEN[base(h['opp'])]['BAAytd'] for h in hit_pop])
sp_baa_z = z([sp_hand_stat(P[h['opp_pid']]['sp'], 'L' if h['bats'] == 'L' else 'R', 'BAA') for h in hit_pop])
hpa = z([(h['hh']['H'] / h['hh']['PA']) if h['hh'] and h['hh']['PA'] else None for h in hit_pop])
rpa = z([(h['hh']['R'] / h['hh']['PA']) if h['hh'] and h['hh']['PA'] else None for h in hit_pop])
rbipa = z([(h['hh']['RBI'] / h['hh']['PA']) if h['hh'] and h['hh']['PA'] else None for h in hit_pop])
for i, h in enumerate(hit_pop):
    p = P[h['opp_pid']]; wsp = p['w_sp']
    quality = .40 * Q['wRC'][i] + .35 * Q['xwOBA'][i] + .25 * Q['ISO'][i]
    power = .30 * Q['BRL'][i] + .25 * Q['HH95'][i] + .25 * Q['xSLG'][i] + .20 * Q['Pull'][i]
    disc = -Q['K'][i]
    platoon = 0.819 * (zsplit[i] - zovr[i]) * wsp
    hitter = .40 * quality + .40 * power + .20 * disc + platoon
    spc = -(.40 * (.55 * comp_z['xFIP'][i] + .45 * comp_z['WHIP'][i]) + .35 * (.40 * comp_z['HH'][i] + .30 * comp_z['HRFB'][i] + .30 * comp_z['FBGB'][i])
            + .25 * (.40 * comp_z['BAA'][i] + .35 * comp_z['BABIP'][i] + .25 * comp_z['OPS'][i]))
    penc = -pen_z[i]
    pitcher = wsp * (spc if p['sp'] else penc) + (1 - wsp) * penc
    h['tb_z'] = hitter * .60 + pitcher * .35 + itt_z[h['team']] * .05
    h['hitter_z'] = hitter; h['pitcher_z'] = pitcher; h['platoon'] = platoon
    # HRR
    hits_z = .6 * hpa[i] + .4 * (wsp * sp_baa_z[i] + (1 - wsp) * pen_baa_z[i])
    runs_z = .6 * rpa[i] + .4 * ttz[h['team']]; rbi_z = .6 * rbipa[i] + .4 * ttz[h['team']]
    h['hrr'] = (.4 * hits_z + .3 * runs_z + .3 * rbi_z) * PA_MULT[h['slot']]
    # market TB O1.5
    eid = gm[h['gp']]['eid']
    q15 = {}
    for mk in ('batter_total_bases', 'batter_total_bases_alternate'):
        for pt, rows in player_quotes(eid, mk, h['name'], 1.5).items(): q15.setdefault(1.5, []).extend(rows)
    ps = price_summary(q15[1.5]) if q15 else None
    h['tb_px'] = ps
    if ps and ps['novig_over'] is not None:
        nv = ps['novig_over'] * 100
        h['tb_prob'] = nv + 2.5 * h['tb_z'] + gm[h['gp']]['tb_wx']
        h['tb_edge'] = h['tb_prob'] - nv
        h['tb_novig'] = nv
        if ps['best_over'] is not None:
            pr = h['tb_prob'] / 100
            h['tb_ev'] = pr * payout(ps['best_over']) - (1 - pr) * 100
        else: h['tb_ev'] = None
    else:
        h['tb_prob'] = h['tb_edge'] = h['tb_ev'] = h['tb_novig'] = None
    # HRR line
    hq = player_quotes(eid, 'batter_hits_runs_rbis', h['name'])
    if hq:
        pt = max(hq, key=lambda k: len(hq[k])); h['hrr_line'] = pt; h['hrr_px'] = price_summary(hq[pt])
    else: h['hrr_line'] = None; h['hrr_px'] = None
    h['park_pp'] = gm[h['gp']]['park_pp']; h['slot_pp'] = SLOT_PP[h['slot']]; h['wx_pp'] = gm[h['gp']]['tb_wx']

# ---------------- ML / TOTAL (algo) ----------------
GAME = {}
def alt_total_price(matchup, point, side):
    d = alt_tot.get(matchup)
    if not d: return None, None
    rows = []
    for bk in d.get('bookmakers', []):
        for mk in bk.get('markets', []):
            if mk['key'] in ('totals', 'alternate_totals'):
                for o in mk['outcomes']:
                    if o.get('point') == point and o['name'] == side: rows.append((bk['key'], o['price']))
    if not rows: return None, None
    med = statistics.median(am_to_prob(r[1]) for r in rows if not is_corrupt_am(r[1]))
    act = [r for r in rows if r[0] in ACTIONABLE]
    best = max(act, key=lambda r: -am_to_prob(r[1]))[1] if act else None
    return prob_to_am(med), best
for g in games:
    gp = str(g['game_pk']); a, h = g['away'], g['home']; gr = gl[gm[gp]['eid']]
    ia, ih = ITT[g['away_k']], ITT[g['home_k']]
    winner = h if ih >= ia else a; margin = abs(ih - ia)
    ml_w = int(gr['med_ml_home'] if winner == h else gr['med_ml_away']); ml_l = int(gr['med_ml_away'] if winner == h else gr['med_ml_home'])
    tot = float(gr['market_total']); proj_t = ia + ih; diff = proj_t - tot
    tside = 'Over' if diff > 0 else 'Under'
    bet_line = tot
    if tot % 1 == 0: bet_line = tot - 0.5 if tside == 'Over' else tot + 0.5
    if bet_line == tot:
        tprice = int(gr['med_total_over'] if tside == 'Over' else gr['med_total_under'])
    else:
        tprice, _ = alt_total_price(gk(a, h, g.get('gn') if g.get('dh') in ('S','Y') else None), bet_line, tside)
    # algo qual
    if ml_w > 0: ml_qual = True; ml_note = 'dog projected winner'
    elif ml_w < -160: ml_qual = False; ml_note = f'{ml_w} outside -160 bar'
    else:
        need = (abs(ml_w) - 100) / 100; ml_qual = margin >= need; ml_note = f'margin {margin:.2f} vs need {need:.2f}'
    # totals qual: whole line -> conservative half-line (the one nearer the model's number is the harder test)
    if tot % 1 == 0:
        test_line = bet_line  # Grant ruling 8/16: whole-number totals test vs the HALF-RUN RUNG (the line actually bet)
        alt_test = tot - 0.5 if tside == 'Under' else tot + 0.5  # harder rung, info only
    else:
        test_line = alt_test = tot
    t_qual = abs(proj_t - test_line) >= 0.75
    t_qual_alt = abs(proj_t - alt_test) >= 0.75
    ma = model_a[g['mkey']]; mb = model_b[g['mkey']]
    GAME[gp] = {'a': a, 'h': h, 'ia': ia, 'ih': ih, 'winner': winner, 'margin': margin, 'ml_w': ml_w, 'ml_l': ml_l, 'tot': tot, 'proj_t': proj_t, 'diff': diff,
                'tside': tside, 'bet_line': bet_line, 'tprice': tprice, 'ml_qual': ml_qual, 'ml_note': ml_note, 't_qual': t_qual, 't_qual_alt': t_qual_alt,
                'test_line': test_line, 'ma': ma, 'mb': mb, 'time': g['time'], 'venue': g['venue'],
                'ml_over': gr['med_total_over'], 'ml_under': gr['med_total_under'], 'ml_home': gr['med_ml_home'], 'ml_away': gr['med_ml_away']}
    # Triple lock check
    locks = []
    # ML: model A >=65 for winner? model B ML rank<=5? algo qual & same team
    a_ml_team, a_ml_p = ma['ml_team'], ma['ml_p']
    if a_ml_p >= 65 and mb['ml_rank'] <= 5 and mb['ml_team'] == a_ml_team and ml_qual and winner == a_ml_team:
        locks.append(f'ML {winner} {ml_w}')
    a_tside, a_tp = ma['tot_side'], ma['tot_p']
    if a_tp >= 65 and mb['tot_rank'] <= 5 and mb['tot_side'] == a_tside and t_qual and tside == a_tside:
        locks.append(f'TOTAL {tside} {bet_line}')
    GAME[gp]['locks'] = locks
    GAME[gp]['near'] = []
    if a_ml_p >= 65 and mb['ml_rank'] <= 5 and mb['ml_team'] == a_ml_team and winner == a_ml_team and not ml_qual: GAME[gp]['near'].append(f'ML {winner}: A/B unanimous, algo {ml_note}')
    if a_tp >= 65 and mb['tot_rank'] <= 5 and mb['tot_side'] == a_tside and tside == a_tside and not t_qual: GAME[gp]['near'].append(f'TOTAL {tside}: A/B unanimous, algo diff {diff:+.2f} vs {test_line} (need 0.75){" — QUALS on the half-run rung " + str(alt_test) if t_qual_alt else ""}')

# ---------------- outputs ----------------
os.makedirs('projections', exist_ok=True); os.makedirs('candidates', exist_ok=True)
FIELDS = ['date', 'game_pk', 'away', 'home', 'market', 'entity', 'player_id', 'line', 'side', 'proj', 'price', 'ts', 'board', 'ev']
rows = []
def R(gp, market, entity, pid='', line='', side='', proj='', price='', board='', ev=''):
    g = GAME[gp]
    rows.append({'date': DATE, 'game_pk': gp, 'away': g['a'], 'home': g['h'], 'market': market, 'entity': entity, 'player_id': pid, 'line': line, 'side': side,
                 'proj': proj, 'price': price if price is not None else '', 'ts': TS, 'board': board, 'ev': ev})
for gp, g in GAME.items():
    R(gp, 'ML', g['winner'], '', '', g['winner'], round(g['margin'], 3), g['ml_w'])
    R(gp, 'TOTAL', f"{g['a']}@{g['h']}", '', g['bet_line'], g['tside'], round(g['proj_t'], 3), g['tprice'])
for t, d in TT.items():
    if d['line'] is not None: R(d['gp'], 'TT', base(t), '', d['line'], d['side'], round(d['score'], 5), d['price'])
for p in KP:
    for mk, ln, px, pr in (('K', p['K_line'], p['K_px'], p['proj_K']), ('BB', p['BB_line'], p['BB_px'], p['proj_BB']), ('ER', p['ER_line'], p['ER_px'], p['proj_ER'])):
        if ln is None: continue
        side = 'Over' if pr > ln else 'Under'
        price = ''
        if px: price = px['med_over'] if side == 'Over' else px['med_under']
        R(p['gp'], mk, p['name'], p['id'], ln, side, round(pr, 4), price)
# TB boards
eligible = [h for h in hit_pop if h['PA'] >= 100 and h['tb_prob'] is not None]
ev_sorted = sorted([h for h in eligible if h['tb_ev'] is not None], key=lambda h: -h['tb_ev'])
ev_board = []; pergame = defaultdict(int)
for h in ev_sorted:
    if pergame[h['gp']] >= 6: continue
    ev_board.append(h); pergame[h['gp']] += 1
    if len(ev_board) == 20: break
z_board = sorted(eligible, key=lambda h: -h['tb_z'])[:20]
evset = {id(h) for h in ev_board}; zset = {id(h) for h in z_board}
tb_rows = {}
for h in ev_board + z_board:
    if id(h) in tb_rows: continue
    tag = 'BOTH' if (id(h) in evset and id(h) in zset) else ('EV' if id(h) in evset else 'Z')
    tb_rows[id(h)] = h
    R(h['gp'], 'TB', h['name'], h['id'], 1.5, 'Over', round(h['tb_z'], 5), h['tb_px']['best_over'] if h['tb_px'] and h['tb_px']['best_over'] is not None else h['tb_px']['med_over'], tag, round(h['tb_ev'], 2) if h['tb_ev'] is not None else '')
# HRR board top 20 by score, max 6/game, PA>=100, lined
hrr_el = sorted([h for h in hit_pop if h['PA'] >= 100 and h['hrr_line'] is not None], key=lambda h: -h['hrr'])
hrr_board = []; pergame = defaultdict(int)
for h in hrr_el:
    if pergame[h['gp']] >= 6: continue
    hrr_board.append(h); pergame[h['gp']] += 1
    if len(hrr_board) == 20: break
for h in hrr_board:
    R(h['gp'], 'HRR', h['name'], h['id'], h['hrr_line'], 'Over', round(h['hrr'], 4), h['hrr_px']['best_over'] if h['hrr_px']['best_over'] is not None else h['hrr_px']['med_over'])
with open(f'projections/{DATE}.csv', 'w', newline='') as f:
    w = csv.DictWriter(f, fieldnames=FIELDS, lineterminator='\r\n'); w.writeheader(); w.writerows(rows)
# candidates
crow = []
tbc = sorted([h for h in hit_pop if h['tb_px'] and h['tb_px']['best_over'] is not None and h['tb_prob'] is not None], key=lambda h: -h['tb_z'])
for i, h in enumerate(tbc):
    crow.append({'date': DATE, 'market': 'TB', 'entity': h['name'], 'player_id': h['id'], 'team': base(h['team']), 'proj': round(h['tb_ev'], 2) if h['tb_ev'] is not None else '', 'z': round(h['tb_z'], 5),
                 'price': h['tb_px']['best_over'], 'rank': i + 1, 'pa': h['PA'], 'on_board': 'Y' if id(h) in tb_rows else 'N'})
hc = sorted([h for h in hit_pop if h['hrr_line'] is not None], key=lambda h: -h['hrr'])
hb = {id(h) for h in hrr_board}
for i, h in enumerate(hc):
    crow.append({'date': DATE, 'market': 'HRR', 'entity': h['name'], 'player_id': h['id'], 'team': base(h['team']), 'proj': round(h['hrr'], 4), 'z': round(h['hrr'], 4),
                 'price': h['hrr_px']['best_over'] if h['hrr_px']['best_over'] is not None else h['hrr_px']['med_over'], 'rank': i + 1, 'pa': h['PA'], 'on_board': 'Y' if id(h) in hb else 'N'})
with open(f'candidates/{DATE}.csv', 'w', newline='') as f:
    w = csv.DictWriter(f, fieldnames=['date', 'market', 'entity', 'player_id', 'team', 'proj', 'z', 'price', 'rank', 'pa', 'on_board'], lineterminator='\r\n'); w.writeheader(); w.writerows(crow)

# board dump for emission
def slim(h): return {k: h[k] for k in ('name', 'team', 'opp', 'gp', 'slot', 'src', 'bats', 'PA', 'tb_z', 'hitter_z', 'pitcher_z', 'platoon', 'tb_prob', 'tb_edge', 'tb_ev', 'tb_novig', 'hrr', 'hrr_line', 'park_pp', 'slot_pp', 'wx_pp', 'no_wb') } | {'best': h['tb_px']['best_over'] if h['tb_px'] else None, 'book': h['tb_px']['best_book'] if h['tb_px'] else None, 'hrr_px': h['hrr_px']}
out = {'GAME': GAME, 'TT': {t: {k: v for k, v in d.items() if k not in ('osp', 'tq', 'px')} | {'osp': d['osp']['name']} for t, d in TT.items()},
       'KP': [{k: p[k] for k in ('name', 'team', 'opp', 'gp', 'hand', 'proj_K', 'proj_BB', 'proj_ER', 'K_line', 'BB_line', 'ER_line', 'K_score', 'BB_score', 'ER_dist', 'outgs_adj', 'w_sp', 'smGS', 'opp_adj_K', 'opp_adj_BB', 'opp_adj_ER', 'recency', 'rec_missing')} | {'K_px': p['K_px'], 'BB_px': p['BB_px'], 'ER_px': p['ER_px'], 'pp': p['pp'], 'algoK': SPRANK.get(norm(p['name']))} for p in KP],
       'ev_board': [slim(h) for h in ev_board], 'z_board': [slim(h) for h in z_board], 'hrr_board': [slim(h) for h in hrr_board],
       'gm': {gp: {k: v for k, v in d.items() if k != 'g'} for gp, d in gm.items()}, 'openers': [p['name'] for p in P.values() if p['opener']],
       'no_wb': sorted({h['name'] for h in hit_pop if h['no_wb']}), 'n_rows': len(rows), 'n_cand': len(crow), 'lg': LG,
       'ev_z_overlap': len(evset & zset)}
json.dump(out, open('board.json', 'w'), indent=1, default=str)
print('rows', len(rows), 'cand', len(crow), 'overlap', len(evset & zset))
from collections import Counter; print(Counter(r['market'] for r in rows))
for gp, g in GAME.items(): print(g['a'], '@', g['h'], 'ML', g['winner'], round(g['margin'], 2), g['ml_w'], g['ml_note'], '|', g['tside'], g['bet_line'], round(g['diff'], 2), 'qual', g['t_qual'], g['tprice'], '| LOCKS', g['locks'], g['near'])
